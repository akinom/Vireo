import argparse
import logging
import openpyxl
import sys
import importlib
# importlib.reload(certificateProgramSplit)



wb = openpyxl.load_workbook(filename = 'ExcelExport.xlsx')
sheet = wb.worksheets[0]
header = next(sheet.iter_rows(min_row=1, max_row=1))

class ArgParser(argparse.ArgumentParser):
    def parse_args(self):
        args= argparse.ArgumentParser.parse_args(self)
        args.wb = openpyxl.load_workbook(filename = args.excel)
        if (1 != len(args.wb.worksheets)) :
            self.error("Workbook '%s' should have exactly one sheet" % (args.excel))
        args.sheet = args.wb.worksheets[0]
        column = 'Certificate Program'
        args.split_col = _col_index_of(args.sheet, column)
        if (None == args.split_col):
            self.error("Workbook '%s' does not contain a '%s' column" % (args.excel, column))
        return args

def _col_index_of(ws, title):
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in  row:
            if (cell.value == title):
                return cell.col_idx -1;
    return None;

def _dump_sheet(ws):
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in  row:
            print("%s" % (cell.value))
        print("---")


def _split_sheet(ws, column):
    splits = {};
    for row in ws.iter_rows(min_row=2):
        col_val = row[column].value
        if (col_val in splits):
            splits[col_val].append(row);
        else:
            splits[col_val] = [row];
    return splits

def _print_tsv(name, header, rows):
    print("> Print %s" % name);
    file_name = name.replace(' ', '-')
    if not file_name:
        file_name = 'None'
    out =open(file_name + ".tsv", 'w')
    _print_row(header, out)
    for r in rows:
        _print_row(r, out)
    out.close();
    print("< Print %s" % name);


def _print_row(row, out):
    #print('\t'.join([c.value for c in row]), file=out)
    print(row, file=sys.stdout)
    print('\t'.join([(c.value if None != c.value else 'None') for c in row]), file=sys.stdout)

def main():
    description = """split vireo excel export file into multiple files based on first column value
    ignore entries that do not have a value in the first column"""
    loglevels = ['CRITICAL', 'ERROR', 'WARN', 'INFO', 'DEBUG', 'NOTSET']

    parser = ArgParser(description=description, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("--excel", "-s", default=None, required=True, help="excel export file from vireo")
    parser.add_argument("--loglevel", choices=loglevels,  default=logging.ERROR, help="log level  - default: ERROR")
    args = parser.parse_args()

    logging.getLogger().setLevel(args.loglevel)
    logging.basicConfig()

    print("opened %s" % args.excel)
    print("opened workbook " % args.wb)
    print("splitting on value in column: %d" % args.split_col)

    splits = _split_sheet(args.sheet, args.split_col);
    header =  next(args.sheet.iter_rows(min_row=1, max_row=1))
    for k in splits.keys():
        _print_tsv(k, header, splits[k])

if __name__ == "__main__":
    main()


name = 'ExcelExport.xlsx'