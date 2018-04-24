import argparse
from argparse import ArgumentError

import logging
import traceback
import openpyxl
import sys
import importlib
from copy import deepcopy;

def interactive():
    importlib.reload(certificateProgramSplit)
    wb = openpyxl.load_workbook(filename = 'ExcelExport.xlsx')
    sheet = wb.worksheets[0]
    header = next(sheet.iter_rows(min_row=1, max_row=1))


class ArgParser(argparse.ArgumentParser):
    def parse_args(self):
        args= argparse.ArgumentParser.parse_args(self)
        try:
            args.vireo = Vireo(args)
        except Exception as e:
            raise e
        return args

class Vireo:
    CERTIFICATE_PROGRAM = 'Certificate Program'
    THESIS_TYPE = 'Thesis Type'
    STUDENT_EMAIL = 'Student email'
    ID = 'ID'

    NOSPLIT_KEY = '___NO_SPLIT___'

    testing = False;

    def __init__(self, args):
        Vireo.testing = (args.loglevel == logging.getLevelName(logging.DEBUG))

        thesis_file = args.thesis
        thesis_wb = openpyxl.load_workbook(filename = thesis_file)
        if (1 != len(thesis_wb.worksheets)) :
            raise Exception("Workbook '%s' should have exactly one sheet" % (thesis_file))
        self.thesis = thesis_wb.worksheets[0]

        self.thesis_id_col = self._col_index_of(self.thesis, Vireo.ID)
        if (None == self.thesis_id_col):
            raise Exception("Workbook '%s' does not contain a '%s' column" % (thesis_file, Vireo.ID))

        if (args.split_col):
            self._split_col = self._col_index_of(self.thesis, args.split_col)
            if (None == self._split_col):
                raise Exception("Workbook '%s' does not contain a '%s' column" % (thesis_file, args.split_col))
        else:
            self._split_col = None

        if (args.add_certs):
            add_certs_file = args.add_certs
            add_certs_wb = openpyxl.load_workbook(filename = add_certs_file)
            if (1 != len(add_certs_wb.worksheets)) :
                raise Exception("Workbook '%s' should have exactly one sheet" % (add_certs_file))
            self.add_certs = add_certs_wb.worksheets[0]

            if (None == self._col_index_of(self.add_certs, Vireo.ID)):
                raise Exception("Workbook '%s' does not contain a '%s' column" % (add_certs_file, Vireo.ID))
            if (None == self._col_index_of(self.add_certs, Vireo.STUDENT_EMAIL)):
                raise Exception("Workbook '%s' does not contain a '%s' column" % (add_certs_file, Vireo.STUDENT_EMAIL))
            if (None == self._col_index_of(self.thesis, Vireo.STUDENT_EMAIL)):
                raise Exception("Workbook '%s' does not contain a '%s' column" % (thesis_file, Vireo.STUDENT_EMAIL))
            if (None == self._col_index_of(self.thesis, Vireo.CERTIFICATE_PROGRAM)):
                raise Exception("Workbook '%s' does not contain a '%s' column" % (thesis_file, Vireo.CERTIFICATE_PROGRAM))

        self.id_rows = {}
        self._id_rows()

        self._add_rows = None

    def log_info(self):
        header = Vireo._header(self.thesis);
        logging.info("thesis workbook %s" % self.thesis.title)
        logging.info("thsis id column: %s (%d)" % (Vireo.ID,  self.thesis_id_col))
        if (self._split_col): logging.info("thsis split column: %s (%d)" % (header[self._split_col].value, self._split_col));
        if ('add_certs' in dir(self)):
            logging.info("thesis check column: %s (%d)" % (Vireo.STUDENT_EMAIL,  self._col_index_of(self.thesis, Vireo.STUDENT_EMAIL)))
            logging.info("thesis certificate column: %s (%d)" % (Vireo.CERTIFICATE_PROGRAM,  self._col_index_of(self.thesis, Vireo.CERTIFICATE_PROGRAM)))
            logging.info("add_certs id column: %s (%d)" % (Vireo.ID,  self._col_index_of(self.add_certs, Vireo.ID)))
            logging.info("add_certs check column: %s (%d)" % (Vireo.STUDENT_EMAIL,  self._col_index_of(self.add_certs, Vireo.STUDENT_EMAIL)))
            logging.info("add_certs certs column: %s (%d)" % (Vireo.CERTIFICATE_PROGRAM,  self._col_index_of(self.add_certs, Vireo.CERTIFICATE_PROGRAM)))

    def main(args):
        vireo = Vireo(args);
        vireo.log_info();
        splits = vireo.split_sheet()
        vireo.print(splits);

    def _col_index_of(self, sheet, title):
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in  row:
                if (cell.value.strip() == title):
                    return cell.col_idx -1;
        return None;

    def split_sheet(self):
        splits = {}
        splits = self._split_rows(splits, self._add_certificates())
        if (Vireo.testing):
            num_rows = sum(1 for _ in  self._add_certificates())
            logging.debug("# %d certificate rows " % num_rows)
            num_split_rows = self.count_rows("add_certs", splits)
            if (num_rows != num_split_rows):
                logging.error("SANITY CHECK FAILED: # cert rows (%d) != #rows in splits (%d)" % (num_rows, num_split_rows))

        splits = self._split_rows(splits, self.thesis.iter_rows(min_row=2))
        if (Vireo.testing):
            num_rows = num_rows + sum(1 for _ in  self.thesis.iter_rows(min_row=2))
            logging.debug("# %d thesis rows + certificate rows " % num_rows)
            num_split_rows = self.count_rows("thesis", splits)
            if (num_rows != num_split_rows):
                logging.error("SANITY CHECK FAILED: #thesis rows (%d) != #rows in splits (%d)" % (num_rows, num_split_rows))

        return splits

    def count_rows(self, label, splits):
        logging.info("%s #%d split_keys " % (label, len(splits)))
        s = 0
        for k in splits:
            s = s + len(splits[k])
            logging.debug("#%s: %d entries in '%s'" % (label, len(splits[k]), str(k)))
        logging.info("%s total entries %d" % (label, s))
        return s

    def _split_rows(self, splits, iter):
        if (None != self._split_col):
            for row in iter:
                col_val = row[self._split_col].value
                if (None != col_val):
                    col_val = col_val.strip()
                else:
                    col_val = '';
                if (col_val in splits):
                    splits[col_val].append(row);
                else:
                    splits[col_val] = [row];
        else:
            for row in iter:
                if (Vireo.NOSPLIT_KEY in splits):
                    splits[Vireo.NOSPLIT_KEY].append(row);
                else:
                    splits[Vireo.NOSPLIT_KEY] = [row];

        return splits

    def _add_certificates(self):
        if (self._add_rows == None):
            add_rows = []
            if ('add_certs' in dir(self)):
                certs_id_col = self._col_index_of(self.add_certs, Vireo.ID);
                certs_cert_col = self._col_index_of(self.add_certs, Vireo.CERTIFICATE_PROGRAM)
                certs_email_col = self._col_index_of(self.add_certs, Vireo.STUDENT_EMAIL)
                thesis_email_col = self._col_index_of(self.thesis, Vireo.STUDENT_EMAIL)
                thesis_cert_col = self._col_index_of(self.thesis, Vireo.CERTIFICATE_PROGRAM)
                if (certs_id_col == None or certs_cert_col == None or certs_email_col == None or thesis_email_col == None or thesis_cert_col == None):
                    raise RuntimeError("programming error: should never get here - see checks in constructor")

                thesis_ids = self._id_rows();
                # go through all data rows in add_certs sheet
                for row in self.add_certs.iter_rows(min_row=2):
                    sub_id = row[self.thesis_id_col].value
                    if (sub_id in thesis_ids):
                        add_cert = deepcopy(thesis_ids[row[self.thesis_id_col].value])
                        if (add_cert[thesis_email_col].value.strip() != row[certs_email_col].value.strip()):
                            logging.error("ERROR: %s and %s disagree on student email for submission with ID %s " % (
                            self.thesis.title, self.add_certs.title, sub_id));
                        else:
                            logging.debug("add_cert %d %s->%s %s" % (sub_id,  add_cert[thesis_cert_col].value, row[thesis_cert_col].value, add_cert[thesis_email_col]))
                            add_cert[thesis_cert_col].value = row[certs_cert_col].value
                            add_rows.append(add_cert)
                    else:
                            logging.error("No thesis with ID %s found in %s" % (sub_id, self.thesis.title))
            self._add_rows = add_rows
        return self._add_rows

    def _id_rows(self):
        if not self.id_rows:
            for row in self.thesis.iter_rows(min_row=2):
                id = row[self.thesis_id_col].value
                if (id in self.id_rows):
                    raise Exception("duplicate id %s" % str(id))
                self.id_rows[id] = row
        return self.id_rows;


    def _header(sheet):
        return next(sheet.iter_rows(min_row=1, max_row=1))

    def print(self, splits):
        logging.info("printing %d tsv files" % len(splits))
        if Vireo.NOSPLIT_KEY in splits:
            logging.info("no splits %s" % Vireo.NOSPLIT_KEY)
            Vireo._print_tsv(Vireo._header(self.thesis), splits[Vireo.NOSPLIT_KEY], sys.stdout)
        else:
            for k in splits:
                self.print_tsv(k, splits[k])

    def print_tsv(self, name, rows):
        file_name = name.replace(' ', '-')
        if not file_name:
            file_name = 'None'
        out =open(file_name + ".tsv", 'w')
        Vireo._print_tsv(Vireo._header(self.thesis), rows, out)
        out.close();

    def _print_tsv(header, rows, out):
        Vireo._print_tsv_row(header, out)
        for r in rows:
            Vireo._print_tsv_row(r, out)


    def _print_tsv_row(row, out):
        print('\t'.join([(str(c.value) if None != c.value else 'None') for c in row]), file=out)

def main():
    description = """read thesis submission info from file given in --thesis option 
if --add_certs option is given read info about additional certificate programs from file                  
               and add appitional entries in the thesis submission list 
if --split_col is given print tsv table of combined submissions to stdout
if --split_col is given print to tsv files, one for each value in thegiven column 
    """
    loglevels = ['CRITICAL', 'ERROR', 'WARN', 'INFO', 'DEBUG', 'NOTSET']

    parser = ArgParser(description=description, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("--thesis", "-t", default=None, required=True, help="excel export file from vireo")
    parser.add_argument("--split_col", "-s", required=False, help="split column name - default %s" % Vireo.CERTIFICATE_PROGRAM)
    parser.add_argument("--add_certs", "-a", default=None, required=False, help="excep spreadsheet with addituion certificate program info")
    parser.add_argument("--loglevel", "-l", choices=loglevels,  default=logging.ERROR, help="log level  - default: ERROR")
    args = parser.parse_args()
    try:
        logging.getLogger().setLevel(args.loglevel)
        logging.basicConfig()

        Vireo.main(args)

    except Exception as e:
        logging.error(e, file = sys.stderr)
        logging.debug(traceback.format_exc())

if __name__ == "__main__":
    main()
