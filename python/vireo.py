import argparse

import logging
import openpyxl

def snippet():
    logging.getLogger().setLevel('INFO')
    # code snippet when in case you want to test in interactive python
    wb = openpyxl.load_workbook(filename = 'Thesis.xlsx')
    sheet = wb.worksheets[0]
    v = VireoSheet('Thesis.xlsx')
    v.log_info()
    certs = v.readMoreCerts('AdditionalPrograms.xlsx')
    certs.log_info()
    return certs


class VireoSheet:
    CERTIFICATE_PROGRAM = 'Certificate Program'
    STUDENT_NAME = 'Student name'
    DEPARTMENT = 'Department'
    SUBMISSION_DATE = 'Submission date'
    ADVISORS = 'Advisors'
    DOCUMENT_TITLE = 'Title'
    PRIMARY_DOCUMENT = 'Primary document'
    THESIS_TYPE = 'Thesis Type'
    STUDENT_EMAIL = 'Student email'
    MULTI_AUTHOR = 'Multi Author'
    STATUS = 'Status'
    TITLE = 'Title'
    ID = 'ID'

    #RESTRICTIONS SHEET
    R_STUDENT_NAME = 'Submitted By'
    R_TITLE = 'Name'
    R_WALK_IN = 'Walk In Access'
    R_EMBARGO = 'Embargo Years'

    def __init__(self, thesis_file, unique_ids=True):
        """
        read from execl file
        check that there is only one sheet with an ID column
        :param thesis_file:   excel file
        """
        self.file_name = thesis_file
        self.unique_ids = unique_ids
        thesis_wb = openpyxl.load_workbook(filename = thesis_file)
        if (1 != len(thesis_wb.worksheets)) :
            raise Exception("%s should have exactly one sheet" % (thesis_file))
        self._sheet = thesis_wb.worksheets[0]
        self.id_col = self.col_index_of(VireoSheet.ID, required=True)
        self.id_rows = {}
        for row in self._sheet.iter_rows(min_row=2):
            if not row[self.id_col].value:
                logging.warning("%s: Row has no ID value: %s" % (self.file_name, str.join(',',(str(cell.value).strip() for cell in row))))
            else:
                id = int(row[self.id_col].value)
                if not id in self.id_rows:
                    self.id_rows[id] = [row]
                elif unique_ids:
                    raise Exception("%s has duplicate id %s" % (thesis_file, str(id)))
                else:
                    self.id_rows[id].append(row)

    def col_names(self):
        hdrs = next(self._sheet.iter_rows(min_row=1, max_row=1))
        return VireoSheet.row_values(hdrs)

    def col_index_of(self, col_header, required=False):
        try:
            return self.col_names().index(col_header)
        except ValueError as e:
            if (required):
                raise Exception("%s does not contain a '%s' column" % (self.file_name, col_header))
        return None;

    staticmethod
    def row_values(row):
        return [str(cell.value).strip() for cell in row]

    def readMoreCerts(self, add_certs_file, check_id=True):
        add_certs = VireoSheet(add_certs_file, unique_ids=False)
        # check that required columns are present
        certs_email_col_id = add_certs.col_index_of(VireoSheet.STUDENT_EMAIL, required=True)
        certs_cert_col_id = add_certs.col_index_of(VireoSheet.CERTIFICATE_PROGRAM, required=True)
        thesis_email_col_id = self.col_index_of(VireoSheet.STUDENT_EMAIL, required=True)
        thesis_cert_col_id = self.col_index_of(VireoSheet.CERTIFICATE_PROGRAM, required=True)
        # look through whether certs file info matches thesis sheet info
        for cert_id in add_certs.id_rows:
            if  not  cert_id in  self.id_rows:
                if (check_id):
                    raise Exception("%s, row with ID %d: there is no such row in %s" % (add_certs.file_name, cert_id, self.file_name))
                else:
                    continue
            thesis_row = self.id_rows[cert_id][0]
            for row in  add_certs.id_rows[cert_id]:
                logging.debug("ID %d -> cert_row: %s" % (cert_id, VireoSheet.row_values(row)))
                # check that student email matches
                if (row[certs_email_col_id].value.strip() != thesis_row[thesis_email_col_id].value.strip()):
                    raise Exception("%s, row with ID %d: email mistmatch with same row in %s" % (add_certs.file_name, cert_id, self.file_name))
                # check that certificate program enry is not empty
                if not row[certs_cert_col_id].value:
                    raise Exception("%s, row with ID %d: row has empty certificate value" % (add_certs.file_name, cert_id))
        return add_certs

    def readRestrictions(self, restriction_file, check_id=True):
        restrictions = VireoSheet(restriction_file, unique_ids=False)
        # check that necessary columns are present
        restrictions.col_index_of(VireoSheet.R_STUDENT_NAME, required=True)
        restrictions.col_index_of(VireoSheet.R_TITLE, required=True)
        restrictions.col_index_of(VireoSheet.R_WALK_IN, required=True)
        restrictions.col_index_of(VireoSheet.R_EMBARGO, required=True)
        # check wheter restriction IDs make sense
        # look through whether certs file info matches thesis sheet info
        for restr_id in restrictions.id_rows:
            if  not  restr_id in  self.id_rows:
                if (check_id):
                    raise Exception("%s, row with ID %d: there is no such row in %s" % (restrictions.file_name, restr_id, self.file_name))
                else:
                    continue
        return restrictions

    def log_info(self):
        logging.info(str(self))
        logging.info("%s: headers %s" % (self.file_name, str.join(',', self.col_names())))
        for col in [VireoSheet.STUDENT_EMAIL, VireoSheet.CERTIFICATE_PROGRAM]:
            logging.info("%s column: %s (%d)" % (self.file_name, col, self.col_index_of(col)))

    def __str__(self):
        return "%s: ID-col:%d, nrows:%d" % (self.file_name, self.id_col, len(self.id_rows))


